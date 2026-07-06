#!/usr/bin/env python3
"""
新卒CA研修「業界×職種マップ」を .xlsx として生成する。
データ正本は gas_ca_training_v1.js（extract_ca_data.js で data_ca_training.json に抽出済み）。
デザインは GAS と同じ Tokumori ブランド（赤 #AF322C ＋黒＋グレー）を openpyxl で再現。
出力した .xlsx は Google ドライブにアップ→「Googleスプレッドシートとして開く」でそのまま使える。
"""

import json
import math
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

# ===== ブランドカラー =====
RED = "AF322C"
BLACK = "000000"
WHITE = "FFFFFF"
INK = "1A1A1A"
SUB = "6B6B6B"
ZEBRA = "F7F5F4"
BORDER = "D9D9D9"
CHIP_MID = "E4E4E4"
CHIP_LOW = "EFEDEC"
DARK = "2B2B2B"

SIDE = Side(style="thin", color=BORDER)
BOX = Border(left=SIDE, right=SIDE, top=SIDE, bottom=SIDE)

BASE = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(BASE, "data_ca_training.json"), encoding="utf-8") as f:
    D = json.load(f)

# タブ名
T = {
    "TOC": "00_使い方",
    "IMAP": "01_業界マップ",
    "IDET": "02_業界詳細",
    "JMAP": "03_職種マップ",
    "JDET": "04_職種詳細",
    "RANK": "05_人気ランキング",
    "TIPS": "06_面談活用メモ",
    "DEEP": "07_深掘りテンプレ",
}


# ===========================================================================
# ヘルパー
# ===========================================================================
def fill(color):
    return PatternFill("solid", fgColor=color)


def colpx(ws, col):
    w = ws.column_dimensions[get_column_letter(col)].width
    return (w or 8.43) * 7.0


def est_lines(text, px):
    per = max(4.0, px / 14.0)  # 全角は概ね 14px/字
    return max(1, math.ceil(len(str(text)) / per))


def set_col_widths(ws, start_col, px_list):
    ws.column_dimensions["A"].width = 24 / 7.0
    for i, px in enumerate(px_list):
        ws.column_dimensions[get_column_letter(start_col + i)].width = round(px / 7.0, 1)


def border_range(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = BOX


def band_title(ws, row, c1, c2, text):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row, c1, text)
    cell.fill = fill(RED)
    cell.font = Font(color=WHITE, bold=True, size=12)
    cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[row].height = 28


def note(ws, row, col, span, text, size=9, color=SUB, bold=False):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    cell = ws.cell(row, col, text)
    cell.font = Font(color=color, size=size, bold=bold)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    spx = sum(colpx(ws, col + x) for x in range(span))
    ws.row_dimensions[row].height = min(120, max(18, est_lines(text, spx) * 15))
    return row + 1


def kpi(ws, label_row, items, start_col, width):
    c = start_col
    for it in items:
        ws.merge_cells(start_row=label_row, start_column=c, end_row=label_row, end_column=c + width - 1)
        lc = ws.cell(label_row, c, it["label"])
        lc.fill = fill(DARK)
        lc.font = Font(color=WHITE, bold=True, size=10)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=label_row + 1, start_column=c, end_row=label_row + 1, end_column=c + width - 1)
        vc = ws.cell(label_row + 1, c, it["value"])
        vc.fill = fill(WHITE)
        vc.font = Font(color=RED if it["accent"] else INK, bold=True, size=24)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        border_range(ws, label_row, c, label_row + 1, c + width - 1)
        c += width
    ws.row_dimensions[label_row].height = 22
    ws.row_dimensions[label_row + 1].height = 44


def table(ws, row, col, header, rows, col_px, red_cols=None, pop_col=None, min_row_h=None):
    w = len(header)
    for j, h in enumerate(header):
        cell = ws.cell(row, col + j, h)
        cell.fill = fill(BLACK)
        cell.font = Font(color=WHITE, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24

    n = len(rows)
    for i, rdata in enumerate(rows):
        rr = row + 1 + i
        bg = ZEBRA if i % 2 == 1 else WHITE
        maxlines = 1
        for j, val in enumerate(rdata):
            cell = ws.cell(rr, col + j, val)
            cell.fill = fill(bg)
            cell.font = Font(color=(RED if (red_cols and j in red_cols) else INK),
                             bold=bool(red_cols and j in red_cols), size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            lines = est_lines(val, col_px[j])
            if lines > maxlines:
                maxlines = lines
        if pop_col is not None:
            cell = ws.cell(rr, col + pop_col)
            v = str(rdata[pop_col])
            if v[:1] == "高":
                cell.fill = fill(RED); cell.font = Font(color=WHITE, bold=True, size=10)
            elif "低" in v:
                cell.fill = fill(CHIP_LOW); cell.font = Font(color=SUB, bold=True, size=10)
            else:
                cell.fill = fill(CHIP_MID); cell.font = Font(color=INK, bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        h = min(180, max(18, maxlines * 15))
        if min_row_h:
            h = max(h, min_row_h)
        ws.row_dimensions[rr].height = h

    border_range(ws, row, col, row + n, col + w - 1)
    return row + n + 2


def merge_table(ws, row, start_col, header, spans, rows, red_cols=None, link_col=None):
    """列ごとに結合幅 spans を指定する表。link_col 指定時はその列をURLハイパーリンク化。"""
    c = start_col
    for k, h in enumerate(header):
        ws.merge_cells(start_row=row, start_column=c, end_row=row, end_column=c + spans[k] - 1)
        cell = ws.cell(row, c, h)
        cell.fill = fill(BLACK)
        cell.font = Font(color=WHITE, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c += spans[k]
    ws.row_dimensions[row].height = 24

    n = len(rows)
    for i, rdata in enumerate(rows):
        rr = row + 1 + i
        c = start_col
        bg = ZEBRA if i % 2 == 1 else WHITE
        maxlines = 1
        for k, val in enumerate(rdata):
            ws.merge_cells(start_row=rr, start_column=c, end_row=rr, end_column=c + spans[k] - 1)
            cell = ws.cell(rr, c, val)
            cell.fill = fill(bg)
            is_red = bool(red_cols and k in red_cols)
            cell.font = Font(color=(RED if is_red else INK), bold=is_red, size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if link_col is not None and k == link_col and str(val).startswith("http"):
                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=str(val), display=str(val))
                cell.font = Font(color="1155CC", size=9, underline="single")
            spx = sum(colpx(ws, c + x) for x in range(spans[k]))
            lines = est_lines(val, spx)
            if lines > maxlines:
                maxlines = lines
            c += spans[k]
        ws.row_dimensions[rr].height = min(190, max(18, maxlines * 15))

    border_range(ws, row, start_col, row + n, start_col + sum(spans) - 1)
    return row + n + 2


def back_link(ws, col, target):
    cell = ws.cell(1, col, "← 目次へ戻る")
    cell.hyperlink = Hyperlink(ref=cell.coordinate, location="'%s'!A1" % target, display="← 目次へ戻る")
    cell.font = Font(color=SUB, size=9, underline="single")
    cell.alignment = Alignment(horizontal="right")
    ws.row_dimensions[1].height = 15


def group_by_dai(items, idx=1):
    order, m = [], {}
    for r in items:
        d = r[idx]
        if d not in m:
            m[d] = []; order.append(d)
        m[d].append(r)
    return [(d, m[d]) for d in order]


# ===========================================================================
# ワークブック・タブ生成
# ===========================================================================
wb = Workbook()
sheets = {}
for i, key in enumerate(["TOC", "IMAP", "IDET", "JMAP", "JDET", "RANK", "TIPS", "DEEP"]):
    ws = wb.active if i == 0 else wb.create_sheet()
    ws.title = T[key]
    ws.sheet_view.showGridLines = False
    sheets[key] = ws


# ---- 00_使い方 ----
def build_toc():
    ws = sheets["TOC"]
    set_col_widths(ws, 2, [96] * 12)
    band_title(ws, 2, 2, 13, "■ 新卒キャリアアドバイザー研修｜業界 × 職種マップ")
    note(ws, 3, 2, 12,
         "面談の前提知識（業界理解・職種理解）を体系的にインプットするための資料です。"
         "学生人気の高い業界から学べる構成。各表はフィルタで絞り込み・並べ替えができます。", size=10, color=INK)
    n_dai = len(group_by_dai(D["INDUSTRIES"]))
    kpi(ws, 5, [
        {"label": "収録 業界数", "value": len(D["INDUSTRIES"]), "accent": True},
        {"label": "収録 職種数", "value": len(D["JOBS"]), "accent": True},
        {"label": "業界 大分類", "value": n_dai, "accent": False},
        {"label": "最終更新", "value": D["UPDATED"], "accent": False},
    ], 2, 3)

    r = 8
    band_title(ws, r, 2, 13, "■ 目次（タブ名をクリックで各シートへ移動）")
    r += 1
    toc = [
        ("IMAP", "10大分類→業界の俯瞰マップ。1業界1行で人気度・代表企業を一覧"),
        ("IDET", "メインのインプット表。各業界をビジネスモデル〜面談ポイントまで詳細に"),
        ("JMAP", "職種の全体像。営業を6軸に分解。大分類→中分類のマップ"),
        ("JDET", "各職種の仕事内容・向き不向き・キャリアパス・面談での伝え方"),
        ("RANK", "最新（2027/2026卒）人気企業ランキング・業界トレンド・就活トレンド"),
        ("TIPS", "面談で押さえる横断ポイントと、学生に説明するための用語集"),
        ("DEEP", "人気業界を1業界1タブで深掘りするための空テンプレ（複製して充填）"),
    ]
    # ヘッダー
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=13)
    for cc, lab in [(2, "タブ"), (4, "内容")]:
        c = ws.cell(r, cc, lab)
        c.fill = fill(BLACK); c.font = Font(color=WHITE, bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 24
    for i, (key, desc) in enumerate(toc):
        rr = r + 1 + i
        bg = ZEBRA if i % 2 == 1 else WHITE
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
        ws.merge_cells(start_row=rr, start_column=4, end_row=rr, end_column=13)
        lc = ws.cell(rr, 2, T[key])
        lc.fill = fill(bg)
        lc.hyperlink = Hyperlink(ref=lc.coordinate, location="'%s'!A1" % T[key], display=T[key])
        lc.font = Font(color="1155CC", bold=True, size=10, underline="single")
        lc.alignment = Alignment(vertical="center")
        dc = ws.cell(rr, 4, desc)
        dc.fill = fill(bg); dc.font = Font(color=INK, size=10)
        dc.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[rr].height = 22
    border_range(ws, r, 2, r + len(toc), 13)
    r = r + len(toc) + 2

    band_title(ws, r, 2, 13, "■ 人気度の凡例")
    r += 1
    for cc, lab, bgc, fc in [(2, "高", RED, WHITE), (5, "中", CHIP_MID, INK), (8, "低", CHIP_LOW, SUB)]:
        ws.merge_cells(start_row=r, start_column=cc, end_row=r, end_column=cc + 2)
        c = ws.cell(r, cc, lab)
        c.fill = fill(bgc); c.font = Font(color=fc, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1
    r = note(ws, r, 2, 12,
             "人気度は各種ランキングを踏まえた体感の目安です。年度・大学・学生の志向で変動します"
             "（断定ではなく面談の入口として活用してください）。", size=9, color=SUB)
    note(ws, r, 2, 12,
         "出典・更新：人気ランキングの数値と出典URLは「05_人気ランキング」タブに記載。データは "
         + D["UPDATED"] + " 時点。", size=9, color=SUB)
    ws.freeze_panes = "A3"


# ---- 01_業界マップ ----
def build_imap():
    ws = sheets["IMAP"]
    px = [160, 330, 70, 300]
    set_col_widths(ws, 2, px)
    back_link(ws, 5, T["TOC"])
    band_title(ws, 2, 2, 5, "■ 業界マップ（大分類 → 中分類の俯瞰）")
    r = note(ws, 3, 2, 4, "人気の高い大分類を上から配置。詳細は「02_業界詳細」へ。", size=9, color=SUB)
    r += 1
    for dai, rows in group_by_dai(D["INDUSTRIES"]):
        band_title(ws, r, 2, 5, "■ %s（%d業界）" % (dai, len(rows)))
        r += 1
        data = [[x[0], x[2], x[8], x[4]] for x in rows]
        r = table(ws, r, 2, ["業界（中分類）", "一言でいうと", "人気度", "代表企業"], data, px, pop_col=2)
    ws.freeze_panes = "D1"


# ---- 02_業界詳細 ----
def build_idet():
    ws = sheets["IDET"]
    px = [130, 95, 200, 220, 175, 175, 195, 220, 64, 270]
    set_col_widths(ws, 2, px)
    back_link(ws, 11, T["TOC"])
    band_title(ws, 2, 2, 11, "■ 業界詳細（面談インプット用マスタ）")
    note(ws, 3, 2, 10, "ヘッダーのフィルタで大分類・人気度などで絞り込み／並べ替えできます。", size=9, color=SUB)
    header = ["業界（中分類）", "大分類", "一言でいうと", "ビジネスモデル・収益源", "代表企業",
              "主要職種", "働き方の特徴", "将来性・トレンド", "人気度", "面談ポイント・よくある誤解"]
    head = 5
    n = len(D["INDUSTRIES"])
    table(ws, head, 2, header, D["INDUSTRIES"], px, pop_col=8)
    ws.freeze_panes = "D6"
    ws.auto_filter.ref = "B5:%s%d" % (get_column_letter(2 + len(header) - 1), head + n)


# ---- 03_職種マップ ----
def build_jmap():
    ws = sheets["JMAP"]
    px = [200, 110, 360]
    set_col_widths(ws, 2, px)
    back_link(ws, 5, T["TOC"])
    band_title(ws, 2, 2, 4, "■ 職種マップ（大分類 → 中分類）")
    r = note(ws, 3, 2, 3, "「営業」は一括りにせず6つの軸で別物として捉えるのがポイント。詳細は「04_職種詳細」へ。",
             size=9, color=SUB)
    r += 1
    band_title(ws, r, 2, 4, "■ 営業を読み解く6軸")
    r += 1
    r = table(ws, r, 2, ["軸", "対比", "ひとことで"], D["SALES_AXES"], px)
    for dai, rows in group_by_dai(D["JOBS"]):
        band_title(ws, r, 2, 4, "■ %s（%d職種）" % (dai, len(rows)))
        r += 1
        data = [[x[0], x[2], x[3]] for x in rows]
        r = table(ws, r, 2, ["職種", "分類軸", "仕事内容（要約）"], data, px)
    ws.freeze_panes = "C1"


# ---- 04_職種詳細 ----
def build_jdet():
    ws = sheets["JDET"]
    px = [160, 110, 80, 220, 160, 150, 150, 175, 175, 120, 260]
    set_col_widths(ws, 2, px)
    back_link(ws, 12, T["TOC"])
    band_title(ws, 2, 2, 12, "■ 職種詳細（面談インプット用マスタ）")
    note(ws, 3, 2, 11, "ヘッダーのフィルタで大分類・分類軸で絞り込み／並べ替えできます。", size=9, color=SUB)
    header = ["職種", "大分類", "分類軸", "仕事内容", "求められる力", "向いている人", "向かない人",
              "代表業界・配属", "キャリアパス", "採用の多さ・人気", "学生の誤解・面談で伝えること"]
    head = 5
    n = len(D["JOBS"])
    table(ws, head, 2, header, D["JOBS"], px)
    ws.freeze_panes = "D6"
    ws.auto_filter.ref = "B5:%s%d" % (get_column_letter(2 + len(header) - 1), head + n)


# ---- 05_人気ランキング ----
def rank_table(ws, row, col, title, data):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    c = ws.cell(row, col, title)
    c.fill = fill(BLACK); c.font = Font(color=WHITE, bold=True, size=10)
    c.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[row].height = 22
    table(ws, row + 1, col, ["#", "企業", "業界"], data, [44, 200, 170], red_cols=[0])
    return row + 1 + len(data)  # 最終データ行


def build_rank():
    ws = sheets["RANK"]
    set_col_widths(ws, 2, [44, 200, 170, 44, 200, 170])
    back_link(ws, 7, T["TOC"])
    band_title(ws, 2, 2, 7, "■ 人気ランキング・トレンド（2027/2026卒）")
    r = 4
    band_title(ws, r, 2, 7, "■ 最初に押さえる要点")
    r += 1
    r = merge_table(ws, r, 2, ["観点", "内容"], [2, 4], D["RANK_NOTES"], red_cols=[0])

    band_title(ws, r, 2, 7, "■ 人気企業ランキング（出典・集計方法で順位は変わる）")
    r += 1
    a = rank_table(ws, r, 2, "マイナビ・日経 27卒｜文系TOP10", D["RANK_MYNAVI_BUNKEI"])
    b = rank_table(ws, r, 5, "マイナビ・日経 27卒｜理系TOP10", D["RANK_MYNAVI_RIKEI"])
    r = max(a, b) + 2
    c = rank_table(ws, r, 2, "ワンキャリア 27卒｜文系TOP10", D["RANK_ONECAREER_B"])
    d = rank_table(ws, r, 5, "ワンキャリア 27卒｜理系TOP10", D["RANK_ONECAREER_R"])
    r = max(c, d) + 2
    e = rank_table(ws, r, 2, "文化放送 27卒｜総合TOP15（ブランド志向）", D["RANK_BUNKA"])
    f = rank_table(ws, r, 5, "学情 27卒｜総合TOP5", D["RANK_GAKUJO"])
    r = max(e, f) + 2

    band_title(ws, r, 2, 7, "■ 業界人気のトレンド")
    r += 1
    r = merge_table(ws, r, 2, ["業界", "動向", "確度"], [2, 3, 1], D["TREND"], red_cols=[0])

    band_title(ws, r, 2, 7, "■ 出典")
    r += 1
    r = merge_table(ws, r, 2, ["調査", "URL"], [2, 4], D["RANK_SOURCES"], link_col=1)
    note(ws, r, 2, 6,
         "注：数値・順位は各調査の集計方法（エントリー実数型／ブランド志向型／お気に入り型）で異なります。"
         "11〜30位や一部業界別順位は最新未確認のため、研修配布時は各URLで補完してください。", size=9, color=SUB)
    ws.freeze_panes = "A3"


# ---- 06_面談活用メモ ----
def build_tips():
    ws = sheets["TIPS"]
    px = [210, 620]
    set_col_widths(ws, 2, px)
    back_link(ws, 3, T["TOC"])
    band_title(ws, 2, 2, 3, "■ 面談で押さえる横断ポイント")
    r = table(ws, 4, 2, ["観点", "内容"], D["TIPS"], px, red_cols=[0])
    band_title(ws, r, 2, 3, "■ 用語集（学生に説明できるように）")
    r += 1
    table(ws, r, 2, ["用語", "意味"], D["GLOSSARY"], px, red_cols=[0])
    ws.freeze_panes = "C1"


# ---- 07_深掘りテンプレ ----
def build_deep():
    ws = sheets["DEEP"]
    px = [320, 520]
    set_col_widths(ws, 2, px)
    back_link(ws, 3, T["TOC"])
    band_title(ws, 2, 2, 3, "■ 業界深掘りテンプレート（このシートを複製して使用）")
    r = note(ws, 3, 2, 2,
             "使い方：このタブを複製し、タブ名を「08_総合商社」等に変更。下表の各章を埋めて、"
             "その業界の『面談で使えるレベル』の深掘り版を作る。まずは人気上位"
             "（総合商社/コンサル/IT/金融/メーカー/広告/人材）から。", size=10, color=INK)
    r += 1
    cell = ws.cell(r, 2, "対象業界：＿＿＿＿＿＿＿＿")
    cell.font = Font(color=RED, bold=True, size=12)
    ws.row_dimensions[r].height = 26
    r += 2
    rows = [[s, ""] for s in D["DEEP_SECTIONS"]]
    table(ws, r, 2, ["章立て", "記入欄"], rows, px, red_cols=[0], min_row_h=52)
    ws.freeze_panes = "C1"


build_toc()
build_imap()
build_idet()
build_jmap()
build_jdet()
build_rank()
build_tips()
build_deep()

OUT_NAME = "新卒CA研修_業界×職種マップ_v1.xlsx"
desktop = os.path.expanduser("~/Desktop")
out_paths = [os.path.join(desktop, OUT_NAME), os.path.join(BASE, OUT_NAME)]
for p in out_paths:
    wb.save(p)
    print("saved:", p)
print("tabs:", [ws.title for ws in wb.worksheets])
